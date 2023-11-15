import requests
from dotenv import load_dotenv
import openpyxl
import datetime
import os

lon = 18.0215
lat = 59.3099

def request_smhi_api():
    """ 
    Hämtar data från SMHI:s API och extraherar värden för temperatur och nederbörd för de kommande 24 timmar för angiven longitude och latitude.

    Returnerar:
        Lista med väderdata i form av listor, varje lista har följande ordningsföljd:
            0. 'created': Skapad tidpunkt ('YYYY-MM-DD hh:mm').
            1. 'longitude': Longitud (decimalgrader).
            2. 'latitude': Latitud (decimalgrader).
            3. 'date': Datum ('YYYY-MM-DD').
            4. 'hour': Timme (0-23).
            5. 'temperature': Temperatur (°C).
            6. 'precipitation': True/False.
            7. 'provider': Dataleverantör ('SMHI').

    Kastar:
        Exception: Vid felaktig statuskod från API-anropet.
    """
    
    response = requests.get(f'https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/{lon}/lat/{lat}/data.json')

    data = []
    
    if response.status_code == 200:
        rounded_start_time = datetime.datetime.now() + datetime.timedelta(minutes = 60 - datetime.datetime.now().minute)
        format_string = "%Y-%m-%d %H:%M"
        formatted_rounded_start_time = rounded_start_time.strftime(format_string)

        for _ in range(24):
            for observation in response.json()['timeSeries']:
                if formatted_rounded_start_time == ' '.join(observation['validTime'].split('T'))[:-4]:
                    created = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    date = formatted_rounded_start_time.split(' ')[0]
                    hour = int(formatted_rounded_start_time.split(' ')[1].split(':')[0])
                    temperature = [param['values'][0] for param in observation['parameters'] if param['name'] == 't'][0]
                    precipitation = [param['values'][0] for param in observation['parameters'] if param['name'] == 'pcat'][0]
                    precipitation_bool = True if precipitation != 0 else False
                    provider = 'SMHI'
                    data.append([created, lon, lat, date, hour, temperature, precipitation_bool, provider])

                    rounded_start_time += datetime.timedelta(hours=1)
                    formatted_rounded_start_time = rounded_start_time.strftime(format_string)
                    break

    else:
        print(f'Felaktig statuskod: {response.status_code}')

    return data


def request_owm_api():
    """ 
    Hämtar data från OpenWeatherMaps API och extraherar värden för temperatur och nederbörd för de kommande 24 timmar för angiven longitude och latitude.

    Returnerar:
        Lista med väderdata i form av listor, varje lista har följande ordningsföljd:
            0. 'created': Skapad tidpunkt ('YYYY-MM-DD hh:mm').
            1. 'longitude': Longitud (decimalgrader).
            2. 'latitude': Latitud (decimalgrader).
            3. 'date': Datum ('YYYY-MM-DD').
            4. 'hour': Timme (0-23).
            5. 'temperature': Temperatur (°C).
            6. 'precipitation': True/False.
            7. 'provider': Dataleverantör ('OpenWeatherMap').

    Kastar:
        Exception: Vid felaktig statuskod från API-anropet.
    """
    load_dotenv()
    API_KEY = os.getenv('API_KEY')
    
    response = requests.get(f'https://api.openweathermap.org/data/3.0/onecall?lat={lat}&lon={lon}&appid={API_KEY}&units=metric&exclude=current,minutely,daily,alerts')
    
    data = []
    
    if response.status_code==200:
        rounded_start_time = datetime.datetime.now() + datetime.timedelta(minutes = 60 - datetime.datetime.now().minute)
        format_string = "%Y-%m-%d %H:%M"
        formatted_rounded_start_time = rounded_start_time.strftime(format_string)

        for _ in range(24):
            for observation in response.json()['hourly']:
                dt_object = datetime.datetime.fromtimestamp(observation['dt'])
                formatted_dt_object = dt_object.strftime('%Y-%m-%d %H:%M')

                if formatted_rounded_start_time == formatted_dt_object:
                    created = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    date = formatted_rounded_start_time.split(' ')[0]
                    hour = int(formatted_rounded_start_time.split(' ')[1].split(':')[0])
                    temperature = float(observation['temp'])
                    precipitation_bool = True if observation['weather'][0]['id'] < 700 else False
                    provider = 'OpenWeatherMap'
                    data.append([created, lon, lat, date, hour, temperature, precipitation_bool, provider])

                    rounded_start_time += datetime.timedelta(hours=1)
                    formatted_rounded_start_time = rounded_start_time.strftime(format_string)
                    break
    else:
        print(f'Felaktig statuskod: {response.status_code}')
    
    return data


def create_excel(data):
    """
    Skapar eller uppdaterar en Excel-fil med väderdata.

    Om filen 'Prognos.xlsx' redan finns, lägger funktionen till nya rader om datan inte redan finns.
    Om filen inte finns, skapas den och fylls med den tillhandahållna datan.

    Args:
        data (list of lists): En lista av listor som innehåller väderdata. Varje inre lista representerar en rad med data.

    Returnerar:
        None

    """
    if os.path.exists('./Prognos.xlsx'):
        workbook = openpyxl.load_workbook('Prognos.xlsx')
        sheet = workbook.active
        
        all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        
        data_needs_append = False
        
        for value in data:
            if value[1:] not in [x[1:] for x in all_rows]:
                data_needs_append = True
                start_row = sheet.max_row+1
                for observation in data:
                    for col_num, value in enumerate(observation, start=1):
                        sheet.cell(row=start_row, column=col_num, value=value)
                    start_row+=1
                print(f'Filen har blivit uppdaterad med senaste data från {"SMHI" if data==request_smhi_api() else "OpenWeatherMap"}.')
                break
        if not data_needs_append:
            print(f'Inga uppdateringar behövs för {"SMHI" if data==request_smhi_api() else "OpenWeatherMap"}.')
        
        workbook.save('Prognos.xlsx')
    
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

        for column_letter in columns:
            sheet.column_dimensions[column_letter].width = 16

        sheet["A1"] = "Created"
        sheet["B1"] = "Longitude"
        sheet["C1"] = "Latitude"
        sheet["D1"] = "Datum"
        sheet["E1"] = "Hour"
        sheet["F1"] = "Temperature"
        sheet["G1"] = "RainOrSnow"
        sheet["H1"] = "Provider"

        for value in data: 
            sheet.append(value)
        workbook.save('Prognos.xlsx')
        print('Filen har blivit skapad.')
        print(f'Filen har blivit uppdaterad med senaste data från {"SMHI" if data==request_smhi_api() else "OpenWeatherMap"}.')

def print_last_update(provider):
    """
    Skriver ut den senaste väderprognosen från en specifik väderleverantör.

    Om filen 'Prognos.xlsx' existerar, hämtar funktionen den senaste väderprognosen från den angivna leverantören
    och skriver ut den i en tabell. Endast de senaste 24 raderna (timmar) av prognoser för den specifika leverantören
    visas.

    Args:
        provider (str): Namnet på väderprognosleverantören.

    Returnerar:
        None

    """
    if os.path.exists('./Prognos.xlsx'):
        try:
            workbook = openpyxl.load_workbook('Prognos.xlsx')
            sheet = workbook.active
            all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            filtered_rows = list(filter(lambda x: x[7] == provider, all_rows))[-24:]

            #print-del:
            title = f'Prognos från {provider} {filtered_rows[0][0]}:'
            print('=' * len(title))
            print(title)
            print('=' * len(title))
            for row in filtered_rows:
                print(f'{datetime.time(row[4]).strftime("%H:%M")} {float(row[5])} grader {"Nederbörd" if row[6]==True else "Ingen nederbörd"}')
        except IndexError:
            print(f'Filen behöver uppdateras med data från {provider}, använd undermeny "1. Hämta data"!')

    else:
        print('Filen existerar inte. Använd undermeny "1. Hämta data" för att skapa filen!')
    
